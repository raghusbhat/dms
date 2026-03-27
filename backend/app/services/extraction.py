import logging
import os
from io import BytesIO

import pdfplumber
import pytesseract
from openpyxl import load_workbook
from pdf2image import convert_from_path
from PIL import Image

logger = logging.getLogger(__name__)


def extract_text(file_path: str, mime_type: str) -> str:
    logger.info("[EXTRACT] Starting — file=%s mime=%s", file_path, mime_type)
    try:
        if mime_type == "application/pdf":
            logger.debug("[EXTRACT] PDF detected — trying pdfplumber")
            text = _extract_from_pdf(file_path)
            logger.debug("[EXTRACT] pdfplumber returned %d chars", len(text))
            if len(text) < 100:
                logger.info("[EXTRACT] Text too short (%d chars) — falling back to OCR", len(text))
                text = _extract_with_ocr(file_path)
                logger.info("[EXTRACT] OCR returned %d chars", len(text))
            return text

        if mime_type in (
            "image/jpeg",
            "image/png",
            "image/tiff",
            "image/bmp",
            "image/webp",
        ):
            logger.debug("[EXTRACT] Image detected — using Tesseract OCR")
            text = _extract_from_image(file_path)
            logger.info("[EXTRACT] Image OCR returned %d chars", len(text))
            return text

        if mime_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ):
            logger.debug("[EXTRACT] Excel detected — using openpyxl")
            text = _extract_from_excel(file_path)
            logger.info("[EXTRACT] Excel returned %d chars", len(text))
            return text

        if mime_type in ("text/plain", "text/csv", "text/markdown", "text/html"):
            logger.debug("[EXTRACT] Plain text detected — reading directly")
            with open(file_path, "r", encoding="utf-8") as f:
                text = f.read()
            logger.info("[EXTRACT] Plain text returned %d chars", len(text))
            return text

        logger.warning("[EXTRACT] Unknown mime type '%s' — treating as PDF", mime_type)
        return _extract_from_pdf(file_path)

    except Exception as e:
        logger.error("[EXTRACT] Failed — file=%s mime=%s error=%s", file_path, mime_type, e, exc_info=True)
        return ""


def _extract_from_pdf(file_path: str) -> str:
    pages_text = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                pages_text.append(text)
    return "\n".join(pages_text)


def _extract_with_ocr(file_path: str) -> str:
    images = convert_from_path(file_path, dpi=200)
    pages_text = []
    for img in images:
        text = pytesseract.image_to_string(img)
        if text:
            pages_text.append(text)
    return "\n".join(pages_text)


def _extract_from_image(file_path: str) -> str:
    with Image.open(file_path) as img:
        return pytesseract.image_to_string(img)


def _extract_from_excel(file_path: str) -> str:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheets_text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows_text = []
        for row in sheet.iter_rows(values_only=True):
            non_empty = [str(cell) if cell is not None else "" for cell in row]
            if any(non_empty):
                rows_text.append(" | ".join(non_empty))
        if rows_text:
            sheets_text.append(f"{sheet_name}\n" + "\n".join(rows_text))
    wb.close()
    return "\n\n".join(sheets_text)
